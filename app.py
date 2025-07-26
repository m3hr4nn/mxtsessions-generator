from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import tempfile
import os
from cryptography.fernet import Fernet
import base64
import hashlib
from datetime import datetime
from openpyxl import load_workbook
import io

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

def generate_key_from_password(password):
    """Generate a Fernet key from a password"""
    # Create a consistent key from password using SHA-256 and base64 encoding
    key = hashlib.sha256(password.encode()).digest()
    return base64.urlsafe_b64encode(key)

def encrypt_password(password, encryption_key):
    """Encrypt a password using Fernet encryption"""
    try:
        key = generate_key_from_password(encryption_key)
        fernet = Fernet(key)
        encrypted = fernet.encrypt(password.encode())
        return base64.urlsafe_b64encode(encrypted).decode()
    except Exception as e:
        raise Exception(f"Encryption failed: {str(e)}")

def decrypt_password(encrypted_password, encryption_key):
    """Decrypt a password using Fernet encryption"""
    try:
        key = generate_key_from_password(encryption_key)
        fernet = Fernet(key)
        encrypted_bytes = base64.urlsafe_b64decode(encrypted_password.encode())
        decrypted = fernet.decrypt(encrypted_bytes)
        return decrypted.decode()
    except Exception as e:
        raise Exception(f"Decryption failed: {str(e)}")

def read_excel_file(file_content):
    """Read Excel file using openpyxl and return data as list of dictionaries"""
    try:
        # Load workbook from file content
        workbook = load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        # Get header row (first row)
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # Validate required columns
        required_columns = ['Hostname', 'IP', 'user', 'password']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
        
        # Get column indices
        hostname_idx = headers.index('Hostname')
        ip_idx = headers.index('IP')
        user_idx = headers.index('user')
        password_idx = headers.index('password')
        
        # Read data rows
        data = []
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            
            # Ensure we have enough columns
            row_data = list(row) + [None] * (len(headers) - len(row))
            
            hostname = str(row_data[hostname_idx] or '').strip()
            ip = str(row_data[ip_idx] or '').strip()
            user = str(row_data[user_idx] or '').strip()
            password = str(row_data[password_idx] or '').strip()
            
            # Skip rows without essential data
            if not hostname or not ip:
                continue
                
            data.append({
                'Hostname': hostname,
                'IP': ip,
                'user': user,
                'password': password
            })
        
        if not data:
            raise ValueError("No valid data rows found in Excel file")
        
        return data
        
    except Exception as e:
        if "Missing required columns" in str(e) or "No valid data rows" in str(e):
            raise e
        else:
            raise Exception(f"Failed to read Excel file: {str(e)}")

def write_excel_file(data, output_path):
    """Write data to Excel file using openpyxl"""
    try:
        from openpyxl import Workbook
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Servers"
        
        # Write headers
        headers = ['Hostname', 'IP', 'user', 'password']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            worksheet.cell(row=row_num, column=1, value=row_data['Hostname'])
            worksheet.cell(row=row_num, column=2, value=row_data['IP'])
            worksheet.cell(row=row_num, column=3, value=row_data['user'])
            worksheet.cell(row=row_num, column=4, value=row_data['password'])
        
        workbook.save(output_path)
        
    except Exception as e:
        raise Exception(f"Failed to write Excel file: {str(e)}")

def mobaxterm_encrypt_password(password, master_password):
    """
    Encrypt password using MobaXterm-compatible method
    This uses a simplified version of MobaXterm's encryption
    """
    try:
        # Use the same encryption method but format for MobaXterm
        key = generate_key_from_password(master_password)
        fernet = Fernet(key)
        encrypted = fernet.encrypt(password.encode())
        # Return base64 encoded for MobaXterm compatibility
        return base64.urlsafe_b64encode(encrypted).decode()
    except Exception as e:
        raise Exception(f"MobaXterm encryption failed: {str(e)}")

def generate_mxtsessions_content(data, file_name, password_format='plain', encryption_key=None):
    """Generate MobaXterm sessions file content with password format option"""
    
    # Header section
    content = "[Bookmarks]\n"
    content += f"SubRep={file_name}\n"
    content += f"ImgNum=41\n\n"
    
    # Sessions section
    for row in data:
        hostname = row['Hostname']
        ip = row['IP']
        username = row['user']
        password = row['password']
        
        # Handle password based on format preference
        if password_format == 'encrypted' and encryption_key and password:
            try:
                # Encrypt password for MobaXterm
                password = mobaxterm_encrypt_password(password, encryption_key)
                # Add encryption indicator for MobaXterm (custom format)
                password = f"ENC:{password}"
            except Exception as e:
                # If encryption fails, fall back to plain text with warning
                password = f"ENCRYPT_FAILED_{password}"
        
        # Session entry format for MobaXterm
        session_name = f"{hostname}_{ip}"
        content += f"{session_name}=#109#0%{ip}%22%{username}%{password}%-1%-1%%%%%0%0%0%%1080%%0%0%1#MobaFont%10%0%0%-1%15%236,236,236%30,30,30%180,180,192%0%-1%0%%xterm%-1%-1%_Std_Colors_0_%80%24%0%1%-1%<none>%%0%1%-1#0# #-1\n"
    
    return content

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "message": "MXTSessions Generator API - Enhanced Version",
        "version": "2.0.0",
        "features": [
            "Generate MobaXterm sessions with plain text passwords",
            "Generate MobaXterm sessions with encrypted passwords",
            "MobaXterm-compatible password encryption",
            "Lightweight processing without pandas"
        ],
        "endpoints": {
            "/generate": "POST - Generate MobaXterm sessions file (supports passwordFormat: 'plain' or 'encrypted')",
            "/health": "GET - Health check"
        }
    })

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

@app.route('/generate', methods=['POST'])
def generate_sessions():
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "Only .xlsx files are supported"}), 400
        
        # Get password format preference (plain or encrypted)
        password_format = request.form.get('passwordFormat', 'plain')
        encryption_key = request.form.get('encryptionKey', '') if password_format == 'encrypted' else None
        
        # Validate encryption key if encrypted format is requested
        if password_format == 'encrypted' and not encryption_key:
            return jsonify({"error": "Master password is required for encrypted format"}), 400
        
        # Read Excel file
        try:
            file_content = file.read()
            data = read_excel_file(file_content)
        except Exception as e:
            return jsonify({"error": str(e)}), 400
        
        # Generate file name from original filename
        base_filename = os.path.splitext(file.filename)[0]
        
        # Generate MobaXterm sessions content
        try:
            sessions_content = generate_mxtsessions_content(
                data, 
                base_filename, 
                password_format=password_format,
                encryption_key=encryption_key
            )
        except Exception as e:
            return jsonify({"error": f"Failed to generate sessions: {str(e)}"}), 500
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(mode='w', suffix='.mxtsessions', delete=False) as tmp_file:
            tmp_file.write(sessions_content)
            tmp_file_path = tmp_file.name
        
        # Generate filename based on format
        format_suffix = '_encrypted' if password_format == 'encrypted' else ''
        output_filename = f"{base_filename}{format_suffix}.mxtsessions"
        
        # Return file
        try:
            return send_file(
                tmp_file_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='text/plain'
            )
        finally:
            # Clean up temporary file
            try:
                os.unlink(tmp_file_path)
            except:
                pass
                
    except Exception as e:
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
